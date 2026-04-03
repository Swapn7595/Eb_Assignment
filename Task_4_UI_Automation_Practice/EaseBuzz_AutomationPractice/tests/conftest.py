# conftest.py
import os
import pytest
import logging
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from utils.config import environments
from utils.config import logger
from utils.config import test_step_logger
from utils.config import TestStepLogger
#from utils.screenshot_manager import capture_failure
from utils.helper import *
import openpyxl
from openpyxl.styles import Font

# Create or load Excel workbook
report_path = "reports/test_report.xlsx"
if os.path.exists(report_path):
    workbook = openpyxl.load_workbook(report_path)
    sheet = workbook.active
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Test Report"
    headers = ["Test Case ID", "Test Title", "Test Steps", "Status", "Error Message", "Execution Time"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

# Track current row for new data
current_row = sheet.max_row + 1

# Initialize TestStepLogger
test_step_logger = TestStepLogger()

# Track existing test case IDs to avoid duplicates
existing_test_ids = set((row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row))

# Counter for test case IDs
test_case_counter = sheet.max_row - 1 if sheet.max_row > 1 else 0


def pytest_addoption(parser):
    parser.addoption(
        "--env", action="store", default="dev", help="Environment to run tests against"
    )
    parser.addoption(
        "--headless", action="store_true", default=False, help="Run browser in headless mode"
    )


@pytest.fixture(scope='session')
def config(request):
    env = request.config.getoption("--env")
    if env in environments:
        return environments[env]
    else:
        raise ValueError(f"Unknown environment: {env}")


@pytest.fixture(scope='session')
def driver(config, request):
    chrome_options = Options()
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--log-level=3')

    # Add headless option if requested
    headless = request.config.getoption('--headless')
    if headless:
        chrome_options.add_argument('--headless=new')  # Use new headless mode for Chrome 109+

    service = Service()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.maximize_window()

    base_url = config['base_url']
    driver.get(base_url)

    yield driver
    driver.quit()


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_protocol(item, nextitem):
    global test_case_counter
    start_time = datetime.now()
    # Reset steps for each test
    if hasattr(test_step_logger, 'steps'):
        test_step_logger.steps.clear()
    else:
        test_step_logger.steps = []
    yield
    end_time = datetime.now()

    # Get test function name and docstring for better title
    test_func = getattr(item, 'function', None)
    test_name = item.name
    doc = test_func.__doc__ if test_func and hasattr(test_func, '__doc__') and test_func.__doc__ else ''
    if doc and doc.strip():
        test_title = doc.strip().split('\n')[0]
    else:
        # Fallback: prettify function name
        test_title = " ".join(word.capitalize() for word in test_name.split("_"))

    test_case_counter += 1
    test_case_id = f"TC_{test_case_counter:03d}"
    test_steps = test_step_logger.get_steps() or "No steps logged"

    item.user_properties.append(("execution_time", str(end_time - start_time)))
    item.user_properties.append(("test_case_id", test_case_id))
    item.user_properties.append(("test_title", test_title))
    item.user_properties.append(("test_steps", test_steps))


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()

    if report.when == "call":
        status = "Passed" if report.outcome == "passed" else "Failed"
        error_message = "" if status == "Passed" else str(report.longrepr) if hasattr(report, 'longrepr') else "Unknown Error"

        # Always fetch from user_properties, fallback to safe values
        user_props = dict(item.user_properties)
        execution_time = user_props.get("execution_time", "N/A")
        test_case_id = user_props.get("test_case_id", f"TC_XXX")
        test_title = user_props.get("test_title", item.name)
        test_steps = user_props.get("test_steps", "No steps logged")

        global current_row
        if (test_case_id, test_title) not in existing_test_ids:
            sheet.cell(row=current_row, column=1, value=test_case_id)
            sheet.cell(row=current_row, column=2, value=test_title)
            sheet.cell(row=current_row, column=3, value=test_steps)
            sheet.cell(row=current_row, column=4, value=status)
            sheet.cell(row=current_row, column=5, value=error_message)
            sheet.cell(row=current_row, column=6, value=execution_time)
            existing_test_ids.add((test_case_id, test_title))
            current_row += 1

    if report.failed:
        driver = item.funcargs.get("driver")
        if driver:
            exc_type = type(call.excinfo.value).__name__ if call.excinfo else "Exception"
            try:
                screenshot_path = capture_failure(driver, item.nodeid, exc_type)
                logger.error(f"[SCREENSHOT SAVED ON FAILURE] {screenshot_path}")
                print(f"\n✓ [SCREENSHOT SAVED ON FAILURE] {screenshot_path}")
            except Exception as e:
                logger.error(f"Failed to capture screenshot on failure: {e}")


@pytest.hookimpl(trylast=True)
def pytest_sessionfinish(session, exitstatus):
    try:
        workbook.save(report_path)
        print(f"Enhanced test report saved to {report_path}")
    except Exception as e:
        print(f"Failed to save test report: {e}")


